import os
from aiogram import types

from magic_filter import F
from openpyxl.styles import Border, Side, Alignment
from openpyxl.workbook import Workbook

from filters.private import IsPrivate
from loader import dp, db


@dp.message_handler(IsPrivate(), F.text == 'Avtomobillar hisoboti', state='*')
async def cars_main_menu(message: types.Message):
    admin = await db.select_admin_sql(telegram_id=message.from_user.id)
    if admin:
        cars_month = await db.select_cars_months()
        cars = Workbook()
        cars_ws = cars.active
        cars_ws.append(['SANA', 'SO\'ROV RAQAMI', 'FOYDALANUVCHI TELEFON RAQAMI', 'USERNAME', 'BRAND', 'MODEL'])

        for car in cars_month:
            user = await db.select_user(telegram_id=car[2])
            if user:
                cars_ws.append([car[1], car[0], user[4], user[3], car[3], car[4]])
            else:
                continue

        cars_max_row = cars_ws.max_row
        cars_start_cell = cars_ws['A1']
        cars_end_cell = cars_ws[f'F{cars_max_row}']
        border_style = Border(left=Side(border_style='thin'),
                              right=Side(border_style='thin'),
                              top=Side(border_style='thin'),
                              bottom=Side(border_style='thin'))
        alignment_style = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for row in cars_ws.iter_rows(min_row=cars_start_cell.row, max_row=cars_end_cell.row,
                                     min_col=cars_start_cell.column, max_col=cars_end_cell.column):
            for cell in row:
                cell.border = border_style
                cell.alignment = alignment_style
        cars.save("Mashinalar hisoboti.xlsx")
        await message.answer_document(
            document=types.InputFile(path_or_bytesio="Mashinalar hisoboti.xlsx")
        )
        os.remove("Mashinalar hisoboti.xlsx")


@dp.message_handler(IsPrivate(), F.text == 'Logistika hisoboti', state='*')
async def main_menu_logistics(message: types.Message):
    admin = await db.select_admin_sql(telegram_id=message.from_user.id)
    if admin:
        logistic_data = await db.select_logistics_months()
        wb = Workbook()
        ws = wb.active
        ws.append(['SANA', 'SO\'ROV RAQAMI', 'FOYDALANUVCHI TELEFON RAQAMI', 'USERNAME', 'HUDUD', 'LOGISTIKA TURI'])
        for n in logistic_data:
            user = await db.select_user(telegram_id=n[2])
            if user:
                ws.append([n[1], n[0], user[4], user[3], n[5], n[6]])
            else:
                continue
        max_row = ws.max_row
        start_cell = ws['A1']
        end_cell = ws[f'F{max_row}']
        border_style = Border(left=Side(border_style='thin'),
                              right=Side(border_style='thin'),
                              top=Side(border_style='thin'),
                              bottom=Side(border_style='thin'))
        alignment_style = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for row in ws.iter_rows(min_row=start_cell.row, max_row=end_cell.row,
                                min_col=start_cell.column, max_col=end_cell.column):
            for cell in row:
                cell.border = border_style
                cell.alignment = alignment_style
        wb.save("Logistika hisoboti.xlsx")
        await message.answer_document(
            document=types.InputFile(path_or_bytesio="Logistika hisoboti.xlsx")
        )
        os.remove("Logistika hisoboti.xlsx")


@dp.message_handler(IsPrivate(), F.text == 'Xaridlar hisoboti', state='*')
async def main_menu(message: types.Message):
    admin = await db.select_admin_sql(telegram_id=message.from_user.id)
    if admin:
        buy_data = await db.select_buy_months()
        wb = Workbook()
        ws = wb.active
        ws.append(['SANA', 'SO\'ROV RAQAMI', 'FOYDALANUVCHI TELEFON RAQAMI', 'USERNAME', 'XARID TURI'])
        for n in buy_data:
            user = await db.select_user(telegram_id=n[2])
            if user:
                ws.append([n[1], n[0], user[4], user[3], n[-1]])
            else:
                continue
        max_row = ws.max_row
        start_cell = ws['A1']
        end_cell = ws[f'E{max_row}']
        border_style = Border(left=Side(border_style='thin'),
                              right=Side(border_style='thin'),
                              top=Side(border_style='thin'),
                              bottom=Side(border_style='thin'))
        alignment_style = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for row in ws.iter_rows(min_row=start_cell.row, max_row=end_cell.row,
                                min_col=start_cell.column, max_col=end_cell.column):
            for cell in row:
                cell.border = border_style
                cell.alignment = alignment_style
        wb.save("Xaridlar hisoboti.xlsx")
        await message.answer_document(
            document=types.InputFile(path_or_bytesio="Xaridlar hisoboti.xlsx")
        )
        os.remove("Xaridlar hisoboti.xlsx")


@dp.message_handler(IsPrivate(), F.text == 'Barchasini yuklab olish', state='*')
async def get_all_datas(message: types.Message):
    all_data = await db.select_all_report()
    wb = Workbook()
    ws = wb.active
    ws.append(['SANA', 'SO\'ROV RAQAMI', 'FOYDALANUVCHI TELEFON RAQAMI', 'USERNAME', 'MASHINA BRENDI',
               'MASHINA MODELI', 'LOGISTIKA HUDUDI', 'LOGISTIKA TURI', 'XARID TURI'])
    for n in all_data:
        user = await db.select_user(telegram_id=n[2])
        if user:
            ws.append([n[1], n[0], user[4], user[3], n[3], n[4], n[5], n[6], n[7]])
        else:
            pass
    max_row = ws.max_row
    start_cell = ws['A1']
    end_cell = ws[f'I{max_row}']
    border_style = Border(left=Side(border_style='thin'),
                          right=Side(border_style='thin'),
                          top=Side(border_style='thin'),
                          bottom=Side(border_style='thin'))
    alignment_style = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in ws.iter_rows(min_row=start_cell.row, max_row=end_cell.row,
                            min_col=start_cell.column, max_col=end_cell.column):
        for cell in row:
            cell.border = border_style
            cell.alignment = alignment_style
    wb.save("Umumiy hisobot.xlsx")
    await message.answer_document(
        document=types.InputFile(path_or_bytesio="Umumiy hisobot.xlsx")
    )
    os.remove("Umumiy hisobot.xlsx")
