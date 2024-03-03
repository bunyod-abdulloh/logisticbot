import os

from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Command
from magic_filter import F
from openpyxl.styles import Border, Side, Alignment
from openpyxl.workbook import Workbook

from filters.private import IsPrivate
from keyboards.default.admin_custom_buttons import admin_users_button
from loader import dp, db, bot
from states.admin_states import Admin


@dp.message_handler(IsPrivate(), F.text == 'Foydalanuvchilar', state='*')
async def view_all_users_menu(message: types.Message):
    await message.answer(
        text='Foydalanuvchilar bo\'limi', reply_markup=admin_users_button
    )


@dp.message_handler(IsPrivate(), F.text == 'Download all users', state='*')
async def download_all_users_menu(message: types.Message):
    users = await db.select_all_users()
    wb = Workbook()
    ws = wb.active
    ws.append(['ID', 'TELEGRAM ID', 'FULL_NAME', 'USER_NAME', 'PHONE_NUMBER'])
    c = 0
    print(users)
    for user in users:
        if user:
            c += 1
            ws.append([user[0], user[1], user[2], user[3], user[4]])
        else:
            continue
    max_row = ws.max_row
    start_cell = ws['A1']
    end_cell = ws[f'E{max_row}']
    border_style = Border(left=Side(border_style='thin'),
                          right=Side(border_style='thin'),
                          top=Side(border_style='thin'),
                          bottom=Side(border_style='thin'))
    alignment_style = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in ws.iter_rows(min_row=start_cell.row, max_row=end_cell.row,
                            min_col=start_cell.column, max_col=end_cell.column):
        for cell in row:
            cell.border = border_style
            cell.alignment = alignment_style
    wb.save("All_users.xlsx")
    await message.answer_document(
        document=types.InputFile(path_or_bytesio="All_users.xlsx"),
        caption=f'Jami {c} ta foydalanuvchi ma\'lumotlari yuklandi!'
    )
    os.remove("All_users.xlsx")


@dp.message_handler(IsPrivate(), F.text == 'Block user', state='*')
async def block_user_menu(message: types.Message):
    await message.answer(
        text='Foydalanuvchi ID raqamini kiriting'
    )
    await Admin.delete_user.set()


@dp.message_handler(state=Admin.delete_user)
async def delete_user_menu(message: types.Message, state: FSMContext):
    await state.finish()
    telegram_id = int(message.text)
    await db.update_user_blocks(
        blocked_user=telegram_id, telegram_none=None, telegram_id=telegram_id
    )
    await message.answer(
        text='Foydalanuvchi bloklandi!'
    )
    try:
        await bot.send_message(
            chat_id=telegram_id,
            text='Sizga botdan foydalanish bo\'yicha cheklov qo\'yildi!',
            reply_markup=types.ReplyKeyboardRemove()
        )
    except Exception:
        pass


@dp.message_handler(IsPrivate(), F.text == 'Unblock user', state='*')
async def edit_blocked_user_menu(message: types.Message):
    await message.answer(
        text='Foydalanuvchi ID raqamini kiriting'
    )
    await Admin.unblock_user.set()


@dp.message_handler(state=Admin.unblock_user)
async def unblock_user_menu(message: types.Message, state: FSMContext):
    await state.finish()
    telegram_id = int(message.text)
    await db.update_user_unblock(
        unblock=None, telegram_id=telegram_id, blocked_user=telegram_id
    )
    await message.answer(
        text='Foydalanuvchi blokdan chiqarildi!'
    )
    try:
        await bot.send_message(
            chat_id=telegram_id,
            text='Sizga botdan qayta foydalanish imkoniyati ochildi! /start buyrug\'ini kiritib botdan '
                 'foydalanishingiz foydalanishingiz mumkin!'
        )
    except Exception:
        pass


@dp.message_handler(IsPrivate(), Command('id'), state='*')
async def get_id_user(message: types.Message, state: FSMContext):
    await state.finish()
    await message.answer(
        text=f'ID: <code>{message.from_user.id}</code>'
    )
